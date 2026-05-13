"""
Bulk Applicability Service
--------------------------
Template generation, Azure blob download + parse, bulk applicability updates.
"""

import asyncio
import io
import logging
from datetime import datetime, timezone

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter
from sqlalchemy import bindparam, func, select, update
from sqlalchemy.ext.asyncio import AsyncSession

from utils.dates import format_date_dmy_month_abbr
from models.documents.bulk_applicability import (
    BulkApplicabilityRequest,
    BulkApplicabilityStatus,
)
from models.documents.document import (
    Document,
    DocumentStatus,
    DocumentType,
    DOCUMENT_TYPE_LABELS,
    ApplicabilityType as DocApplicabilityType,
)
from models.events.event import (
    Event,
    EventStatus,
    ApplicabilityType as EventApplicabilityType,
)
from models.events.user import User
from storage import get_storage

logger = logging.getLogger(__name__)

ALLOWED_UPLOAD_EXTENSIONS = frozenset({"xlsx", "csv", "xls"})

# Row 1: A–D empty, then organization_vertical above each division column (not a data column).
# Row 2: id, type, name, updated_at, then division_cluster headers.
FIXED_COLUMNS = ["id", "type", "name", "updated_at"]

# Safety cap so one upload cannot grow unbounded memory in parsed row list.
MAX_BULK_APPLICABILITY_DATA_ROWS = 100_000

# Hard cap on physical data rows scanned (including blank rows) to limit CPU/DoS.
MAX_BULK_APPLICABILITY_SCAN_ROWS = 500_000

# Fewer round-trips than one UPDATE per row; keeps bind parameter batches bounded.
_BULK_UPDATE_CHUNK = 500

_BULK_APP_VALUES = frozenset({"ALL", "DIVISION"})
_DOC_APPLICABILITY_FOR_FILTER = (DocApplicabilityType.ALL, DocApplicabilityType.DIVISION)
_EVT_APPLICABILITY_FOR_FILTER = (EventApplicabilityType.ALL, EventApplicabilityType.DIVISION)


def _is_division_applicability(app_type: object) -> bool:
    v = getattr(app_type, "value", app_type)
    return v == "DIVISION"


def _format_error_for_storage(exc: BaseException) -> str:
    """User-readable text stored on BulkApplicabilityRequest.error_message."""
    if isinstance(exc, ValueError) and "Validation errors" in str(exc):
        return (
            "The uploaded file has invalid or missing data. Ensure header row "
            "includes 'id' and 'type', and each data row has values for 'id' and "
            "'type' (division columns must be Y or N).\n\n"
            f"{exc}"
        )
    if isinstance(exc, ValueError):
        return f"File format or content error:\n{exc}"
    return str(exc)


async def generate_template(
    db: AsyncSession,
    *,
    mode: str,
    selected_types: list[str] | None = None,
    document_ids: list[int] | None = None,
    event_ids: list[int] | None = None,
) -> tuple[io.BytesIO, str]:
    """Build an Excel template in memory and return (bytes_buffer, filename).

    Modes:
      - ``ALL``: include every active record for each type in ``selected_types``.
      - ``SPECIFIC``: include only the documents/events whose ids are listed.
    """
    division_columns = await _get_division_columns(db)
    rows: list[dict] = []

    if (mode or "").upper() == "SPECIFIC":
        if document_ids:
            rows.extend(await _fetch_documents_by_ids(db, document_ids))
        if event_ids:
            rows.extend(await _fetch_events_by_ids(db, event_ids))
        filename = _specific_filename(document_ids, event_ids)
    else:
        types = selected_types or []
        doc_types = [t for t in types if t != "EVENTS"]
        if doc_types:
            rows.extend(await _fetch_active_document_rows(db, doc_types))
        if "EVENTS" in types:
            rows.extend(await _fetch_active_event_rows(db))
        filename = _all_filename(types)

    wb = Workbook()
    ws = wb.active
    ws.title = "Applicability"

    _write_reference_row(ws, division_columns)
    _write_header_row(ws, division_columns)
    _write_data_rows(ws, rows, division_columns)
    _apply_styles(ws, division_columns, len(rows))

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf, filename


def _all_filename(selected_types: list[str]) -> str:
    type_label = "_".join(t.lower() for t in selected_types[:3])
    if len(selected_types) > 3:
        type_label += f"_and_{len(selected_types) - 3}_more"
    return f"bulk_applicability_template_{type_label or 'all'}.xlsx"


def _specific_filename(document_ids: list[int] | None, event_ids: list[int] | None) -> str:
    parts: list[str] = []
    if document_ids:
        parts.append(f"docs{len(document_ids)}")
    if event_ids:
        parts.append(f"events{len(event_ids)}")
    label = "_".join(parts) or "specific"
    return f"bulk_applicability_template_{label}.xlsx"


async def create_upload_request(
    db: AsyncSession,
    *,
    file_name: str,
    blob_path: str,
    selected_types: list[str],
    change_remarks: str | None,
    user_id: str,
) -> BulkApplicabilityRequest:
    req = BulkApplicabilityRequest(
        file_name=file_name,
        uploaded_file_url=blob_path,
        selected_types=selected_types,
        status=BulkApplicabilityStatus.PENDING,
        change_remarks=change_remarks,
        created_by=user_id,
        updated_by=user_id,
    )
    db.add(req)
    await db.flush()
    return req




async def get_history_by_id(
    db: AsyncSession,
    request_id: int,
) -> BulkApplicabilityRequest | None:
    result = await db.execute(
        select(BulkApplicabilityRequest).where(BulkApplicabilityRequest.id == request_id)
    )
    return result.scalar_one_or_none()


async def list_history(
    db: AsyncSession,
    page: int = 1,
    page_size: int = 10,
) -> tuple[list[BulkApplicabilityRequest], int]:
    count_q = select(func.count()).select_from(BulkApplicabilityRequest)
    total = (await db.execute(count_q)).scalar_one()

    offset = (page - 1) * page_size
    items_q = (
        select(BulkApplicabilityRequest)
        .order_by(BulkApplicabilityRequest.created_at.desc())
        .offset(offset)
        .limit(page_size)
    )
    result = await db.execute(items_q)
    items = list(result.scalars().all())
    return items, total


async def process_request_by_id(
    db: AsyncSession,
    request_id: int,
    user_id: str | None = None,
) -> dict:
    """
    Process a single PENDING bulk applicability row (e.g. right after upload).
    Returns final status and optional error_message from the row.
    """
    result = await db.execute(
        select(BulkApplicabilityRequest).where(BulkApplicabilityRequest.id == request_id)
    )
    req = result.scalar_one_or_none()
    if req is None:
        raise ValueError(f"Bulk applicability request {request_id} not found")
    if req.status != BulkApplicabilityStatus.PENDING:
        return {
            "request_id": request_id,
            "status": req.status.value,
            "error_message": req.error_message,
            "message": "Request was not pending; skipped processing.",
        }
    await _process_single_request(db, req, user_id)
    row = await get_history_by_id(db, request_id)
    if row is None:
        raise ValueError(f"Bulk applicability request {request_id} not found after processing")
    return {
        "request_id": request_id,
        "status": row.status.value,
        "error_message": row.error_message,
    }


async def _persist_request_failed(
    db: AsyncSession,
    req_id: int,
    actor: str,
    exc: BaseException,
) -> None:
    """Set FAILED + error_message and commit (caller must not have rolled back the request row)."""
    err_text = _format_error_for_storage(exc)
    await db.execute(
        update(BulkApplicabilityRequest)
        .where(BulkApplicabilityRequest.id == req_id)
        .values(
            status=BulkApplicabilityStatus.FAILED,
            error_message=err_text,
            updated_by=actor,
        )
    )
    await db.flush()
    await db.commit()
    if isinstance(exc, ValueError):
        logger.warning(
            "Bulk applicability request %s failed validation: %s",
            req_id,
            err_text[:500],
        )
    else:
        logger.exception("Bulk applicability request %s failed: %s", req_id, exc)


async def _process_single_request(
    db: AsyncSession,
    req: BulkApplicabilityRequest,
    user_id: str | None = None,
) -> None:
    req_id = req.id
    actor = user_id or req.updated_by or req.created_by
    updater = user_id or req.created_by

    await db.execute(
        update(BulkApplicabilityRequest)
        .where(BulkApplicabilityRequest.id == req_id)
        .values(
            status=BulkApplicabilityStatus.IN_PROGRESS,
            updated_by=actor,
        )
    )
    await db.flush()

    # Parse first. Do NOT session.rollback() here: if upload+process share one HTTP
    # transaction, rollback would undo the new bulk_applicability row so UPDATE … FAILED
    # would match 0 rows.
    try:
        storage = get_storage()
        file_bytes = await storage.read_bytes(req.uploaded_file_url.strip())
        parsed_rows = await asyncio.to_thread(
            _parse_uploaded_file_from_bytes, file_bytes, req.file_name
        )
    except Exception as exc:
        await _persist_request_failed(db, req_id, actor, exc)
        return

    try:
        await _apply_bulk_updates(db, parsed_rows, updater)
    except Exception as exc:
        # Revert partial document/event updates only.
        await db.rollback()
        await _persist_request_failed(db, req_id, actor, exc)
        return

    await db.execute(
        update(BulkApplicabilityRequest)
        .where(BulkApplicabilityRequest.id == req_id)
        .values(
            status=BulkApplicabilityStatus.COMPLETED,
            error_message=None,
            updated_by=actor,
        )
    )
    await db.flush()
    await db.commit()


def _cell_str(value: object) -> str:
    if value is None:
        return ""
    return str(value)


def _prepare_indices(
    reference_row: list[str],
    header_row: list[str],
) -> tuple[int, int, list[tuple[int, str, str]]]:
    header = [h.strip() for h in header_row]
    id_idx = _find_column(header, "id")
    type_idx = _find_column(header, "type")
    updated_at_idx = _find_column(header, "updated_at")
    division_indices: list[tuple[int, str, str]] = []
    for i, h in enumerate(header):
        if i <= updated_at_idx:
            continue
        organization_vertical = reference_row[i].strip() if i < len(reference_row) else ""
        division_indices.append((i, h, organization_vertical))
    if not division_indices:
        raise ValueError("No division columns found in the uploaded file")
    return id_idx, type_idx, division_indices


def _append_parsed_row(
    row: list[str],
    row_num: int,
    id_idx: int,
    type_idx: int,
    division_indices: list[tuple[int, str, str]],
    errors: list[str],
    parsed: list[dict],
) -> None:
    if not any(cell.strip() for cell in row):
        return

    row_id = row[id_idx].strip() if id_idx < len(row) else ""
    row_type = row[type_idx].strip() if type_idx < len(row) else ""

    if not row_id:
        errors.append(f"Row {row_num}: missing 'id'")
        return

    try:
        int(row_id)
    except ValueError:
        errors.append(f"Row {row_num}: 'id' must be an integer, got '{row_id}'")
        return

    if not row_type:
        errors.append(f"Row {row_num}: missing 'type'")
        return

    matched_divisions: set[str] = set()
    has_any_y_or_n: bool = False
    for col_idx, division_name, _organization_vertical in division_indices:
        val = row[col_idx].strip().upper() if col_idx < len(row) else ""
        if val in ("Y", "YES", "N", "NO"):
            has_any_y_or_n = True
        if val in ("Y", "YES"):
            matched_divisions.add(division_name)
        elif val in ("N", "NO", ""):
            pass
        else:
            errors.append(
                f"Row {row_num}, column '{division_name}': "
                f"invalid value '{row[col_idx].strip()}'. Expected Y or N."
            )

    # No division cell was filled with Y/N: leave this document/event unchanged.
    if not has_any_y_or_n:
        return

    parsed.append({
        "id": int(row_id),
        "type": row_type,
        "divisions": sorted(matched_divisions),
        "row_num": row_num,
    })


def _raise_if_errors(errors: list[str]) -> None:
    if errors:
        lines = "\n".join(errors)
        raise ValueError(
            "Validation errors (fix each row/column, then re-upload):\n" + lines
        )


def _parse_uploaded_file_from_bytes(data: bytes, file_name: str) -> list[dict]:
    """Parse file body downloaded from Azure. CSV uses StringIO; Excel uses in-memory workbook."""
    ext = file_name.rsplit(".", 1)[-1].lower() if "." in file_name else ""
    if ext == "csv":
        return _parse_csv_bytes(data)
    if ext in ("xlsx", "xls"):
        return _parse_excel_from_bytes(data)
    raise ValueError(f"Unsupported file extension: {ext}")


def _parse_csv_bytes(data: bytes) -> list[dict]:
    import csv

    errors: list[str] = []
    parsed: list[dict] = []
    text = data.decode("utf-8-sig")
    reader = csv.reader(io.StringIO(text))
    try:
        ref = next(reader)
        hdr = next(reader)
    except StopIteration as exc:
        raise ValueError(
            "CSV must have reference row, header row, and at least one data row"
        ) from exc
    reference_row = [str(c) if c is not None else "" for c in ref]
    header_row = [str(c) if c is not None else "" for c in hdr]
    id_idx, type_idx, division_indices = _prepare_indices(reference_row, header_row)
    for row_num, raw in enumerate(reader, start=1):
        if row_num > MAX_BULK_APPLICABILITY_SCAN_ROWS:
            raise ValueError(
                f"Too many rows in file (max {MAX_BULK_APPLICABILITY_SCAN_ROWS} data rows)."
            )
        row = [str(c) if c is not None else "" for c in raw]
        _append_parsed_row(row, row_num, id_idx, type_idx, division_indices, errors, parsed)
        if len(parsed) > MAX_BULK_APPLICABILITY_DATA_ROWS:
            raise ValueError(
                f"Too many non-empty data rows (max {MAX_BULK_APPLICABILITY_DATA_ROWS})."
            )
    _raise_if_errors(errors)
    return parsed


def _parse_excel_from_bytes(file_bytes: bytes) -> list[dict]:
    wb = load_workbook(io.BytesIO(file_bytes), read_only=True, data_only=True)
    try:
        return _consume_excel_rows(wb.active)
    finally:
        wb.close()


def _consume_excel_rows(ws) -> list[dict]:
    it = ws.iter_rows(values_only=True)
    try:
        r0 = next(it)
        r1 = next(it)
    except StopIteration as exc:
        raise ValueError(
            "Excel must have reference row, header row, and at least one data row"
        ) from exc
    reference_row = [_cell_str(c) for c in r0]
    header_row = [_cell_str(c) for c in r1]
    id_idx, type_idx, division_indices = _prepare_indices(reference_row, header_row)
    errors: list[str] = []
    parsed: list[dict] = []
    for row_num, raw in enumerate(it, start=1):
        if row_num > MAX_BULK_APPLICABILITY_SCAN_ROWS:
            raise ValueError(
                f"Too many rows in file (max {MAX_BULK_APPLICABILITY_SCAN_ROWS} data rows)."
            )
        row = [_cell_str(c) for c in raw] if raw else []
        _append_parsed_row(row, row_num, id_idx, type_idx, division_indices, errors, parsed)
        if len(parsed) > MAX_BULK_APPLICABILITY_DATA_ROWS:
            raise ValueError(
                f"Too many non-empty data rows (max {MAX_BULK_APPLICABILITY_DATA_ROWS})."
            )
    _raise_if_errors(errors)
    return parsed


def _find_column(header: list[str], name: str) -> int:
    for i, h in enumerate(header):
        if h.lower() == name.lower():
            return i
    raise ValueError(
        f"Required column '{name}' is missing from row 2 (header row). "
        f"Use the downloaded template; found columns: {header}"
    )



async def _get_division_columns(db: AsyncSession) -> list[tuple[str, str]]:
    """
    Distinct (organization_vertical, division_cluster) from users.
    Template row 1: org vertical only above division columns (cols after updated_at); row 2: headers.
    """
    result = await db.execute(
        select(User.organization_vertical, User.division_cluster)
        .where(User.division_cluster.isnot(None))
        .where(User.division_cluster != "")
        .distinct()
        .order_by(User.organization_vertical, User.division_cluster)
    )
    return [(row[0] or "", row[1]) for row in result.all()]


def _document_row(r) -> dict:
    refs = r.applicability_refs
    if refs is not None:
        refs = list(refs)
    return {
        "id": r.id,
        "type": DOCUMENT_TYPE_LABELS.get(r.document_type, r.document_type.value),
        "name": r.name,
        "updated_at": r.updated_at,
        "applicability_type": r.applicability_type,
        "applicability_refs": refs,
    }


def _event_row(r) -> dict:
    refs = r.applicability_refs
    if refs is not None:
        refs = list(refs)
    return {
        "id": r.id,
        "type": "Events",
        "name": r.event_name,
        "updated_at": r.updated_at,
        "applicability_type": r.applicability_type,
        "applicability_refs": refs,
    }


async def _fetch_active_document_rows(
    db: AsyncSession, doc_types: list[str]
) -> list[dict]:
    type_enums = [DocumentType(t) for t in doc_types]
    result = await db.execute(
        select(
            Document.id,
            Document.name,
            Document.document_type,
            Document.updated_at,
            Document.applicability_type,
            Document.applicability_refs,
        )
        .where(
            Document.document_type.in_(type_enums),
            Document.status == DocumentStatus.ACTIVE,
            Document.replaces_document_id.is_(None),
            Document.applicability_type.in_(_DOC_APPLICABILITY_FOR_FILTER),
        )
        .order_by(Document.document_type, Document.id)
    )
    return [_document_row(r) for r in result.all()]


async def _fetch_active_event_rows(db: AsyncSession) -> list[dict]:
    result = await db.execute(
        select(
            Event.id,
            Event.event_name,
            Event.updated_at,
            Event.applicability_type,
            Event.applicability_refs,
        )
        .where(
            Event.status == EventStatus.ACTIVE,
            Event.replaces_document_id.is_(None),
            Event.applicability_type.in_(_EVT_APPLICABILITY_FOR_FILTER),
        )
        .order_by(Event.id)
    )
    return [_event_row(r) for r in result.all()]


async def _fetch_documents_by_ids(
    db: AsyncSession, document_ids: list[int]
) -> list[dict]:
    result = await db.execute(
        select(
            Document.id,
            Document.name,
            Document.document_type,
            Document.updated_at,
            Document.applicability_type,
            Document.applicability_refs,
        )
        .where(
            Document.id.in_(document_ids),
            Document.applicability_type.in_(_DOC_APPLICABILITY_FOR_FILTER),
        )
        .order_by(Document.document_type, Document.id)
    )
    return [_document_row(r) for r in result.all()]


async def _fetch_events_by_ids(
    db: AsyncSession, event_ids: list[int]
) -> list[dict]:
    result = await db.execute(
        select(
            Event.id,
            Event.event_name,
            Event.updated_at,
            Event.applicability_type,
            Event.applicability_refs,
        )
        .where(
            Event.id.in_(event_ids),
            Event.applicability_type.in_(_EVT_APPLICABILITY_FOR_FILTER),
        )
        .order_by(Event.id)
    )
    return [_event_row(r) for r in result.all()]


def _write_reference_row(
    ws,
    division_columns: list[tuple[str, str]],
) -> None:
    """Row 1: fixed columns empty; each division column shows organization_vertical (reference)."""
    for c in range(1, len(FIXED_COLUMNS) + 1):
        ws.cell(row=1, column=c, value="")
    for col_idx, (organization_vertical, _division_cluster) in enumerate(
        division_columns, start=len(FIXED_COLUMNS) + 1
    ):
        ws.cell(row=1, column=col_idx, value=organization_vertical)


def _write_header_row(
    ws,
    division_columns: list[tuple[str, str]],
) -> None:
    """Row 2: visible header row."""
    headers = FIXED_COLUMNS + [division_cluster for _vertical, division_cluster in division_columns]
    for col_idx, header in enumerate(headers, start=1):
        ws.cell(row=2, column=col_idx, value=header)


def _write_data_rows(
    ws,
    rows: list[dict],
    division_columns: list[tuple[str, str]],
) -> None:
    """Row 3+: id, type, name, updated_at; then division Y/N grid."""
    for row_idx, row_data in enumerate(rows, start=3):
        ws.cell(row=row_idx, column=1, value=row_data["id"])
        ws.cell(row=row_idx, column=2, value=row_data["type"])
        ws.cell(row=row_idx, column=3, value=row_data["name"])
        updated = row_data.get("updated_at")

        if isinstance(updated, datetime):
            ws.cell(row=row_idx, column=4, value=format_date_dmy_month_abbr(updated))
        else:
            ws.cell(row=row_idx, column=4, value=str(updated) if updated else "")

        refs_raw = row_data.get("applicability_refs") or []
        ref_set = {str(r).strip() for r in refs_raw if r is not None and str(r).strip()}
        is_div = _is_division_applicability(row_data.get("applicability_type"))

        for i, (_vertical, division_cluster) in enumerate(division_columns):
            col = len(FIXED_COLUMNS) + 1 + i
            cell_val = ""
            if is_div and division_cluster in ref_set:
                cell_val = "Y"
            ws.cell(row=row_idx, column=col, value=cell_val)


def _apply_styles(
    ws,
    division_columns: list[tuple[str, str]],
    data_row_count: int,
) -> None:
    header_font = Font(bold=True, size=11)
    header_fill = PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid")
    header_alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    total_cols = len(FIXED_COLUMNS) + len(division_columns)

    # Row 1: organization_vertical reference above divisions.
    reference_fill = PatternFill(start_color="F2F2F2", end_color="F2F2F2", fill_type="solid")
    for col_idx in range(1, total_cols + 1):
        ref_cell = ws.cell(row=1, column=col_idx)
        ref_cell.font = Font(bold=True, size=10)
        ref_cell.alignment = (
            header_alignment if col_idx > len(FIXED_COLUMNS) else Alignment(horizontal="left", vertical="center")
        )
        ref_cell.fill = reference_fill

    for col_idx in range(1, total_cols + 1):
        cell = ws.cell(row=2, column=col_idx)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment

    for col_idx in range(1, total_cols + 1):
        col_letter = get_column_letter(col_idx)
        if col_idx <= len(FIXED_COLUMNS):
            ws.column_dimensions[col_letter].width = 20
        else:
            ws.column_dimensions[col_letter].width = 14

    ws.freeze_panes = "A3"
    ws.sheet_properties.tabColor = "1F4E79"
    ws.protection.sheet = False


# ---------------------------------------------------------------------------
# Upload handling
# ---------------------------------------------------------------------------


# ---------------------------------------------------------------------------
# DB bulk update
# ---------------------------------------------------------------------------

_TYPE_LABEL_TO_DOC_TYPE: dict[str, str] = {}
for dt in DocumentType:
    label = DOCUMENT_TYPE_LABELS.get(dt, dt.value)
    _TYPE_LABEL_TO_DOC_TYPE[label.upper()] = dt.value
    _TYPE_LABEL_TO_DOC_TYPE[dt.value.upper()] = dt.value


def _resolve_type(raw_type: str) -> tuple[str, str]:
    """
    Return (table, enum_value).
    table is 'document' or 'event'.
    """
    upper = raw_type.strip().upper()
    if upper in ("EVENT", "EVENTS"):
        return ("event", "EVENTS")
    resolved = _TYPE_LABEL_TO_DOC_TYPE.get(upper)
    if resolved:
        return ("document", resolved)
    raise ValueError(f"Unknown type '{raw_type}'")


async def _ensure_bulk_targets_all_or_division(
    db: AsyncSession,
    doc_updates: list[dict],
    event_updates: list[dict],
) -> None:
    """Reject EMPLOYEE (or anything other than ALL/DIVISION) for targeted rows."""
    if doc_updates:
        ids = list({u["id"] for u in doc_updates})
        r = await db.execute(
            select(Document.id, Document.applicability_type).where(Document.id.in_(ids))
        )
        bad = [
            row.id
            for row in r.all()
            if row.applicability_type.value not in _BULK_APP_VALUES
        ]
        if bad:
            raise ValueError(
                "Bulk applicability only updates records with applicability ALL or DIVISION. "
                f"Unsupported document IDs: {sorted(bad)}"
            )
    if event_updates:
        ids = list({u["id"] for u in event_updates})
        r = await db.execute(
            select(Event.id, Event.applicability_type).where(Event.id.in_(ids))
        )
        bad = [
            row.id
            for row in r.all()
            if row.applicability_type.value not in _BULK_APP_VALUES
        ]
        if bad:
            raise ValueError(
                "Bulk applicability only updates records with applicability ALL or DIVISION. "
                f"Unsupported event IDs: {sorted(bad)}"
            )


async def _apply_bulk_updates(
    db: AsyncSession,
    rows: list[dict],
    user_id: str,
) -> None:
    """
    Validate IDs exist, then batch UPDATE.
    Rows only include records where at least one division cell was Y/N; others were skipped at parse time.
    """
    doc_updates: list[dict] = []
    event_updates: list[dict] = []

    for row in rows:
        table, _ = _resolve_type(row["type"])
        divisions = row["divisions"]

        if divisions:
            app_type = "DIVISION"
            app_refs = list(divisions)
        else:
            app_type = "ALL"
            app_refs = None

        entry = {
            "id": row["id"],
            "applicability_type": app_type,
            "applicability_refs": app_refs,
            "row_num": row["row_num"],
        }

        if table == "document":
            doc_updates.append(entry)
        else:
            event_updates.append(entry)

    if doc_updates:
        await _validate_ids_exist(db, Document, [u["id"] for u in doc_updates], "document")
    if event_updates:
        await _validate_ids_exist(db, Event, [u["id"] for u in event_updates], "event")
    if doc_updates or event_updates:
        await _ensure_bulk_targets_all_or_division(db, doc_updates, event_updates)

    if doc_updates:
        await _batch_update_documents(db, doc_updates, user_id)
    if event_updates:
        await _batch_update_events(db, event_updates, user_id)


async def _validate_ids_exist(
    db: AsyncSession,
    model,
    ids: list[int],
    label: str,
) -> None:
    result = await db.execute(select(model.id).where(model.id.in_(ids)))
    found = {r[0] for r in result.all()}
    missing = set(ids) - found
    if missing:
        raise ValueError(
            f"The following {label} IDs do not exist: {sorted(missing)}"
        )


async def _batch_update_documents(
    db: AsyncSession,
    updates: list[dict],
    user_id: str,
) -> None:
    # Core Table update (not ORM entity) so executemany + bindparam works without
    # SQLAlchemy's "bulk UPDATE by primary key" parameter naming rules.
    tbl = Document.__table__
    now = datetime.now(timezone.utc)
    stmt = (
        update(tbl)
        .where(tbl.c.id == bindparam("doc_id"))
        .values(
            applicability_type=bindparam("app_type"),
            applicability_refs=bindparam("app_refs"),
            updated_by=bindparam("upd_by"),
            updated_at=bindparam("upd_at"),
        )
    )
    for i in range(0, len(updates), _BULK_UPDATE_CHUNK):
        chunk = updates[i : i + _BULK_UPDATE_CHUNK]
        await db.execute(
            stmt,
            [
                {
                    "doc_id": entry["id"],
                    "app_type": DocApplicabilityType(entry["applicability_type"]),
                    "app_refs": entry["applicability_refs"],
                    "upd_by": user_id,
                    "upd_at": now,
                }
                for entry in chunk
            ],
        )


async def _batch_update_events(
    db: AsyncSession,
    updates: list[dict],
    user_id: str,
) -> None:
    tbl = Event.__table__
    now = datetime.now(timezone.utc)
    stmt = (
        update(tbl)
        .where(tbl.c.id == bindparam("evt_id"))
        .values(
            applicability_type=bindparam("app_type"),
            applicability_refs=bindparam("app_refs"),
            updated_by=bindparam("upd_by"),
            updated_at=bindparam("upd_at"),
        )
    )
    for i in range(0, len(updates), _BULK_UPDATE_CHUNK):
        chunk = updates[i : i + _BULK_UPDATE_CHUNK]
        await db.execute(
            stmt,
            [
                {
                    "evt_id": entry["id"],
                    "app_type": EventApplicabilityType(entry["applicability_type"]),
                    "app_refs": entry["applicability_refs"],
                    "upd_by": user_id,
                    "upd_at": now,
                }
                for entry in chunk
            ],
        )

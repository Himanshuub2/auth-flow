"""
FAQ Service
------------
Fetches the active FAQ document from blob storage, parses the Excel file
(Section / Question / Response), and returns a dict grouped by section.
Results are cached and invalidated when FAQ documents are created or updated.
"""

import asyncio
import io
import logging

from fastapi import HTTPException, status
from openpyxl import load_workbook
from openpyxl.utils.exceptions import InvalidFileException
from sqlalchemy import select
from sqlalchemy.ext.asyncio import AsyncSession

from cache import cache_get, cache_set
from models.documents.document import Document, DocumentStatus, DocumentType
from models.documents.document_file import DocumentFile
from storage import get_storage
from utils import cache_keys

logger = logging.getLogger(__name__)

FAQ_CACHE_TTL_SECONDS = 2 * 24 * 60 * 60  # 2 days

REQUIRED_HEADERS = ("section", "question", "response")


def _normalize_header(value: object | None) -> str:
    if value is None:
        return ""
    return str(value).strip().lower()


def _header_indices(header_row: tuple) -> tuple[int, int, int]:
    """Map Section / Question / Response columns; raise 404 if any are missing."""
    headers = [_normalize_header(cell) for cell in header_row]
    indices: dict[str, int] = {}
    for name in REQUIRED_HEADERS:
        try:
            indices[name] = headers.index(name)
        except ValueError:
            raise HTTPException(
                status_code=status.HTTP_404_NOT_FOUND,
                detail=(
                    f"FAQ file is missing required column '{name.title()}'. "
                    "Expected headers: Section, Question, Response"
                ),
            )
    return indices["section"], indices["question"], indices["response"]


def _parse_faq_excel(file_bytes: bytes) -> dict[str, dict[str, str]]:
    """Parse FAQ xlsx bytes into {section: {question: response}}."""
    if not file_bytes:
        raise HTTPException(
            status_code=status.HTTP_404_NOT_FOUND,
            detail="FAQ file not found or is empty",
        )

    try:
        wb = load_workbook(filename=io.BytesIO(file_bytes), read_only=True, data_only=True)
    except (InvalidFileException, OSError, ValueError) as exc:
        logger.warning("FAQ workbook could not be read: %s", exc)
        raise HTTPException(
            status_code=status.HTTP_404_NOT_FOUND,
            detail="FAQ file not found or could not be read",
        ) from exc

    try:
        ws = wb.active
        if ws is None:
            raise HTTPException(
                status_code=status.HTTP_404_NOT_FOUND,
                detail="FAQ file has no worksheet",
            )

        header_row = next(ws.iter_rows(min_row=1, max_row=1, values_only=True), None)
        if not header_row:
            raise HTTPException(
                status_code=status.HTTP_404_NOT_FOUND,
                detail="FAQ file is missing header row",
            )

        section_idx, question_idx, response_idx = _header_indices(header_row)
        max_col = max(section_idx, question_idx, response_idx)
        result: dict[str, dict[str, str]] = {}

        for row in ws.iter_rows(min_row=2, values_only=True):
            if not row or len(row) <= max_col:
                continue
            section, question, response = row[section_idx], row[question_idx], row[response_idx]
            section_str = str(section).strip() if section else ""
            question_str = str(question).strip() if question else ""
            response_str = str(response).strip() if response else ""
            if not section_str or not question_str or not response_str:
                continue
            result.setdefault(section_str, {})[question_str] = response_str

        if not result:
            raise HTTPException(
                status_code=status.HTTP_404_NOT_FOUND,
                detail="FAQ file has no valid question and answer rows",
            )
        return result
    finally:
        wb.close()


async def get_faq_data(db: AsyncSession) -> dict[str, dict[str, str]]:
    """Return parsed FAQ data, using cache when available."""
    key = cache_keys.faq_data()
    cached = await cache_get(key)
    if cached is not None:
        return cached

    stmt = (
        select(DocumentFile.file_url)
        .join(Document, DocumentFile.document_id == Document.id)
        .where(
            Document.document_type == DocumentType.FAQ,
            Document.status == DocumentStatus.ACTIVE,
        )
        .order_by(Document.updated_at.desc(), DocumentFile.sort_order.asc())
        .limit(1)
    )
    blob_path = (await db.execute(stmt)).scalar_one_or_none()
    if not blob_path:
        raise HTTPException(
            status_code=status.HTTP_404_NOT_FOUND,
            detail="Active FAQ not found",
        )

    storage = get_storage()
    try:
        file_bytes = await storage.read_bytes(blob_path)
    except OSError as exc:
        logger.warning("FAQ blob download failed for %s: %s", blob_path, exc)
        raise HTTPException(
            status_code=status.HTTP_404_NOT_FOUND,
            detail="FAQ file not found",
        ) from exc

    faq_data = await asyncio.to_thread(_parse_faq_excel, file_bytes)

    await cache_set(key, faq_data, ttl=FAQ_CACHE_TTL_SECONDS)
    return faq_data
